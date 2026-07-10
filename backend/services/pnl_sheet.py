"""Импорт помесячных затрат P&L из Google-таблицы финансовой модели.

Таблица открыта «по ссылке», поэтому тянем её **публичный xlsx-экспорт**
(`.../export?format=xlsx`) обычным httpx — без сервис-аккаунта и OAuth. Лист «P&L»:
колонка B — метрика, колонки-месяцы (G…) подписаны диапазоном дат в строке 2
(«01.06 - 30.06»); порядок месяцев в таблице нестандартный (Авг перед Июл), поэтому
месяц берём из подписи диапазона, а не из позиции. Год выводим из сезона (декабрь —
прошлый год относительно `today`).

Мапим только ₽-суммы постоянных статей и ставки-% в `PnlMonth`. Операционный ФОТ НЕ
импортируем — он считается из графика смен (`schedule.labor_for_period`); списания/
упаковка живут на дневном уровне (`PnlDayCost`), из таблицы берём лишь помесячную
упаковку как резерв. Админ-ФОТ (управляющий) — ОТДЕЛЬНАЯ строка «Admin Labor ₽» →
поле `labor_admin`; «Прочие ₽» → `other_opex` (IT/ОФД/эквайринг/аморт.) их НЕ включает,
это непересекающиеся статьи — не склеиваем.
"""

import io
import re
import warnings

import httpx
from openpyxl import load_workbook

from config import settings
from models import PnlMonth, SessionLocal
from utils import today

_EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx"
_RANGE_RE = re.compile(r"(\d{2})\.(\d{2})")

# нормализованная подпись метрики (startswith) → поле PnlMonth. Админ-ФОТ и «Прочие» —
# ОТДЕЛЬНЫЕ поля (чтобы в P&L было видно, сколько съедает управленческий ФОТ).
_MONEY_MAP = [
    ("аренда ₽", "rent"),
    ("коммуналка ₽", "utilities"),
    ("маркетинг ₽", "marketing"),
    ("прочие ₽", "other_opex"),
    ("admin labor ₽", "labor_admin"),
    ("непредвиденные расходы ₽", "contingency"),
    ("кап-резерв ₽", "cap_reserve"),
    ("упаковка ₽", "packaging"),
]
# ставки: доля в таблице (0.06) → проценты в PnlMonth (6)
_PCT_MAP = [
    ("налог %", "tax_pct"),
    ("удержания агрегатора", "aggregator_pct"),
]


def _norm(s: object) -> str:
    """Нижний регистр + схлопнутые пробелы (в таблице встречается двойной пробел)."""
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _num(v: object) -> float:
    """Число из ячейки. Формулы-ошибки (#DIV/0!) и пустые → 0."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\xa0", " ").replace(" ", "").replace("%", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _month_columns(ws) -> dict[int, tuple[int, int]]:
    """{индекс колонки openpyxl → (год, месяц)} из подписей-диапазонов в строке 2."""
    base_year = today().year
    out: dict[int, tuple[int, int]] = {}
    for col in range(7, ws.max_column + 1):  # G и правее
        label = ws.cell(row=2, column=col).value
        m = _RANGE_RE.search(str(label or ""))
        if not m:
            continue
        month = int(m.group(2))
        # сезон Дек(прошлый год)…Авг(текущий): месяцы 9–12 относятся к прошлому году
        year = base_year - 1 if month >= 9 else base_year
        out[col] = (year, month)
    return out


def import_pnl_sheet() -> dict:
    """Скачать таблицу, распарсить лист «P&L», заполнить `PnlMonth`. Идемпотентно.

    Возвращает сводку: сколько месяцев импортировано/пропущено и их значения. Месяцы
    без данных по аренде (напр. текущий/будущий) пропускаются — чтобы не обнулять
    введённое вручную.
    """
    url = _EXPORT_URL.format(sid=settings.pnl_sheet_id)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        resp = httpx.get(url, follow_redirects=True, timeout=30.0)
        resp.raise_for_status()
        wb = load_workbook(io.BytesIO(resp.content), data_only=True)
    ws = wb["P&L"] if "P&L" in wb.sheetnames else wb[wb.sheetnames[0]]

    cols = _month_columns(ws)
    # (год, месяц) → {поле: ₽/%}
    data: dict[tuple[int, int], dict] = {ym: {} for ym in cols.values()}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=False):
        label = _norm(row[1].value)  # колонка B
        if not label:
            continue
        money_field = next((f for key, f in _MONEY_MAP if label.startswith(key)), None)
        pct_field = next((f for key, f in _PCT_MAP if label.startswith(key)), None)
        if not money_field and not pct_field:
            continue
        for col, ym in cols.items():
            val = _num(row[col - 1].value)
            bucket = data[ym]
            if money_field:
                bucket[money_field] = bucket.get(money_field, 0.0) + val
            elif pct_field:
                bucket[pct_field] = val * 100

    imported, skipped = [], []
    with SessionLocal() as db:
        for (year, month), vals in sorted(data.items()):
            # месяц без аренды — данных нет (текущий/будущий), не трогаем
            if _num(vals.get("rent")) <= 0:
                skipped.append({"year": year, "month": month})
                continue
            obj = db.query(PnlMonth).filter_by(year=year, month=month).one_or_none()
            if obj is None:
                obj = PnlMonth(year=year, month=month)
                db.add(obj)
            for field in (
                "rent",
                "utilities",
                "marketing",
                "labor_admin",
                "other_opex",
                "contingency",
                "cap_reserve",
                "packaging",
            ):
                setattr(obj, field, round(vals.get(field, 0.0), 2))
            if "tax_pct" in vals:
                obj.tax_pct = round(vals["tax_pct"], 2)
            if "aggregator_pct" in vals:
                obj.aggregator_pct = round(vals["aggregator_pct"], 2)
            obj.work_hours = obj.work_hours or 12
            imported.append(
                {
                    "year": year,
                    "month": month,
                    **{
                        k: round(vals.get(k, 0.0), 2)
                        for k in (
                            "rent",
                            "utilities",
                            "marketing",
                            "labor_admin",
                            "other_opex",
                            "contingency",
                            "cap_reserve",
                            "packaging",
                            "tax_pct",
                            "aggregator_pct",
                        )
                    },
                }
            )
        db.commit()

    return {"imported": imported, "skipped": skipped, "months": len(imported)}
