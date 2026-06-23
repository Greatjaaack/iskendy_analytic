"""Роутер поставщиков: CRUD, карточка с товарами/ценами, загрузка/скачивание файлов."""

from io import BytesIO

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, Response
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, field_validator
from sqlalchemy import func, select

import storage
from models import Ingredient, SessionLocal, Supplier, SupplierContact, SupplierFile, SupplierPrice
from utils import normalize_email, normalize_phone

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(prefix="/api/suppliers", tags=["suppliers"])


class SupplierIn(BaseModel):
    name: str
    address: str = ""
    min_delivery: str = ""
    comment: str = ""


class ContactIn(BaseModel):
    contact_person: str = ""
    phone: str = ""
    whatsapp: str = ""
    telegram: str = ""
    email: str = ""
    comment: str = ""

    @field_validator("phone", "whatsapp")
    @classmethod
    def _valid_phone(cls, v: str) -> str:
        v = (v or "").strip()
        return normalize_phone(v) if v else ""

    @field_validator("email")
    @classmethod
    def _valid_email(cls, v: str) -> str:
        return normalize_email(v)

    @field_validator("telegram")
    @classmethod
    def _clean_telegram(cls, v: str) -> str:
        return (v or "").strip()


def _contact_dict(c: SupplierContact) -> dict:
    return {
        "id": c.id,
        "contact_person": c.contact_person,
        "phone": c.phone,
        "whatsapp": c.whatsapp,
        "telegram": c.telegram,
        "email": c.email,
        "comment": c.comment,
    }


def _supplier_brief(s: Supplier, products: int = 0) -> dict:
    return {
        "id": s.id,
        "name": s.name,
        "address": s.address,
        "min_delivery": s.min_delivery,
        "comment": s.comment,
        "contacts": [_contact_dict(c) for c in s.contacts],
        "products": products,
    }


@router.get("")
def list_suppliers():
    with SessionLocal() as db:
        counts = dict(
            db.execute(
                select(
                    SupplierPrice.supplier_id,
                    func.count(func.distinct(SupplierPrice.ingredient_id)),
                ).group_by(SupplierPrice.supplier_id)
            ).all()
        )
        rows = db.execute(select(Supplier).order_by(Supplier.name)).scalars().all()
        return [_supplier_brief(s, counts.get(s.id, 0)) for s in rows]


def _contact_line(c: SupplierContact) -> str:
    """Однострочное представление контакта для Excel."""
    parts = []
    if c.contact_person:
        parts.append(c.contact_person)
    if c.phone:
        parts.append(f"тел: {c.phone}")
    if c.whatsapp:
        parts.append(f"WA: {c.whatsapp}")
    if c.telegram:
        parts.append(f"TG: {c.telegram}")
    if c.email:
        parts.append(f"email: {c.email}")
    if c.comment:
        parts.append(f"({c.comment})")
    return " · ".join(parts)


@router.get("/export")
def export_suppliers():
    """Выгрузка всех поставщиков с контактами/каналами в .xlsx."""
    with SessionLocal() as db:
        counts = dict(
            db.execute(
                select(
                    SupplierPrice.supplier_id,
                    func.count(func.distinct(SupplierPrice.ingredient_id)),
                ).group_by(SupplierPrice.supplier_id)
            ).all()
        )
        rows = db.execute(select(Supplier).order_by(Supplier.name)).scalars().all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Поставщики"
        headers = [
            "Поставщик",
            "Контакты и каналы",
            "Адрес",
            "Мин. поставка",
            "Комментарий",
            "Товаров",
        ]
        ws.append(headers)
        for s in rows:
            ws.append(
                [
                    s.name,
                    "\n".join(_contact_line(c) for c in s.contacts),
                    s.address,
                    s.min_delivery,
                    s.comment,
                    counts.get(s.id, 0),
                ]
            )
        widths = [28, 50, 30, 18, 30, 10]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        buf = BytesIO()
        wb.save(buf)

    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": 'attachment; filename="suppliers.xlsx"'},
    )


@router.get("/{supplier_id}")
def get_supplier(supplier_id: int):
    with SessionLocal() as db:
        s = db.get(Supplier, supplier_id)
        if not s:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "поставщик не найден")

        # товары с последней ценой
        prices = db.execute(
            select(SupplierPrice, Ingredient)
            .join(Ingredient, SupplierPrice.ingredient_id == Ingredient.id)
            .where(SupplierPrice.supplier_id == supplier_id)
            .order_by(Ingredient.name, SupplierPrice.price_date)
        ).all()
        latest: dict[int, dict] = {}
        for p, ing in prices:
            latest[ing.id] = {
                "ingredient_id": ing.id,
                "price_id": p.id,
                "name": ing.name,
                "brand": p.brand,
                "unit": ing.unit,
                "pack_size": p.pack_size,
                "pack_price": p.pack_price,
                "unit_price": p.unit_price,
                "price_date": p.price_date.isoformat() if p.price_date else None,
            }

        files = [
            {
                "id": f.id,
                "filename": f.filename,
                "file_type": f.file_type,
                "uploaded_at": f.uploaded_at.isoformat(),
            }
            for f in sorted(s.files, key=lambda x: x.uploaded_at, reverse=True)
        ]

        return {
            **_supplier_brief(s, len(latest)),
            "products_list": sorted(latest.values(), key=lambda x: x["name"]),
            "files": files,
        }


@router.post("")
def create_supplier(data: SupplierIn):
    with SessionLocal() as db:
        if db.execute(select(Supplier).where(Supplier.name == data.name)).scalar_one_or_none():
            raise HTTPException(status.HTTP_409_CONFLICT, "поставщик с таким именем уже есть")
        s = Supplier(**data.model_dump())
        db.add(s)
        db.commit()
        db.refresh(s)
        return _supplier_brief(s)


@router.put("/{supplier_id}")
def update_supplier(supplier_id: int, data: SupplierIn):
    with SessionLocal() as db:
        s = db.get(Supplier, supplier_id)
        if not s:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "поставщик не найден")
        for k, v in data.model_dump().items():
            setattr(s, k, v)
        db.commit()
        return _supplier_brief(s)


class ProductPatch(BaseModel):
    brand: str = ""


@router.delete("/{supplier_id}")
def delete_supplier(supplier_id: int):
    """Удалить поставщика со всеми контактами, ценами и файлами (каскад + файлы с диска)."""
    with SessionLocal() as db:
        s = db.get(Supplier, supplier_id)
        if not s:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "поставщик не найден")
        # физические файлы удаляем до ORM-каскада (после commit пути уже не достать)
        for f in s.files:
            storage.delete_file(f.path)
        db.delete(s)  # cascade="all, delete-orphan" уберёт контакты/цены/файлы из БД
        db.commit()
        return {"ok": True, "id": supplier_id}


@router.put("/{supplier_id}/products/{price_id}")
def update_product(supplier_id: int, price_id: int, data: ProductPatch):
    """Обновить торговую марку конкретной позиции прайса поставщика."""
    with SessionLocal() as db:
        p = db.get(SupplierPrice, price_id)
        if not p or p.supplier_id != supplier_id:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "позиция не найдена")
        p.brand = data.brand.strip()
        db.commit()
        return {"id": p.id, "brand": p.brand}


@router.post("/{supplier_id}/contacts")
def add_contact(supplier_id: int, data: ContactIn):
    with SessionLocal() as db:
        if not db.get(Supplier, supplier_id):
            raise HTTPException(status.HTTP_404_NOT_FOUND, "поставщик не найден")
        c = SupplierContact(supplier_id=supplier_id, **data.model_dump())
        db.add(c)
        db.commit()
        db.refresh(c)
        return _contact_dict(c)


@router.put("/{supplier_id}/contacts/{contact_id}")
def update_contact(supplier_id: int, contact_id: int, data: ContactIn):
    with SessionLocal() as db:
        c = db.get(SupplierContact, contact_id)
        if not c or c.supplier_id != supplier_id:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "контакт не найден")
        for k, v in data.model_dump().items():
            setattr(c, k, v)
        db.commit()
        db.refresh(c)
        return _contact_dict(c)


@router.delete("/{supplier_id}/contacts/{contact_id}")
def delete_contact(supplier_id: int, contact_id: int):
    with SessionLocal() as db:
        c = db.get(SupplierContact, contact_id)
        if not c or c.supplier_id != supplier_id:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "контакт не найден")
        db.delete(c)
        db.commit()
        return {"status": "ok"}


@router.post("/{supplier_id}/files")
async def upload_file(
    supplier_id: int, file: UploadFile = File(...), file_type: str = Form("other")
):
    with SessionLocal() as db:
        s = db.get(Supplier, supplier_id)
        if not s:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "поставщик не найден")
        data = await file.read()
        path, original = storage.save_bytes(data, file.filename or "file")
        rec = SupplierFile(
            supplier_id=supplier_id, filename=original, path=path, file_type=file_type
        )
        db.add(rec)
        db.commit()
        db.refresh(rec)
        return {
            "id": rec.id,
            "filename": rec.filename,
            "file_type": rec.file_type,
            "uploaded_at": rec.uploaded_at.isoformat(),
        }


@router.get("/{supplier_id}/files/{file_id}")
def download_file(supplier_id: int, file_id: int):
    with SessionLocal() as db:
        f = db.get(SupplierFile, file_id)
        if not f or f.supplier_id != supplier_id:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "файл не найден")
        return FileResponse(f.path, filename=f.filename)
