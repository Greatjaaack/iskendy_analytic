"""
Импорт файла «ТТК и матрица продуктов.xlsx» в БД.

Структура файла (68 листов):
  - «Матрица поставщиков» — прайс: Ингредиент | Поставщик | Фасовка | Цена упаковки |
    Цена за 1 ед.изм. | Дата проценки | (хвост — заметки).
  - «Матрица продуктов» — Ингредиент | Ед.изм. | Вкусвилл | Рынок-Бухта | turkish-food.
  - «Сводная» — категории (Дюрюмы/Напитки/…) и блюда.
  - По листу на каждую ТТК: B1 = имя; строка-заголовок «Ингредиент,Брутто,…»;
    строки ингредиентов до «Итого» (в «Итого» col7=себестоимость, col9=выход).
    Полуфабрикаты — листы с префиксом «пф …», в строках зовутся «п/ф …».

Себестоимость берём готовую из файла (колонка «с/с руб.»), а не пересчитываем.
Связь строк ТТК с сырьём/п/ф — по нормализованному имени.
"""

import logging
import os
import re
from datetime import date, datetime

import openpyxl

from models import (
    DishMapping,
    Ingredient,
    SessionLocal,
    Supplier,
    SupplierPrice,
    SyncLog,
    Ttk,
    TtkIngredient,
)

logger = logging.getLogger(__name__)

SEED_PATH = os.path.join(os.path.dirname(__file__), "..", "seed", "ttk_matrix.xlsx")

SERVICE_SHEETS = {"Матрица продуктов", "Матрица поставщиков", "Сводная", "Доставка"}
PRODUCT_SUPPLIER_COLS = {3: "Вкусвилл", 4: "Рынок-Бухта", 5: "turkish-food"}


_PF_PREFIX = re.compile(r"^\s*п\s*[/\\]?\s*ф\s+", re.I)


def _strip_pf(s) -> str:
    return _PF_PREFIX.sub("", str(s or "").strip())


def _is_pf_ref(s) -> bool:
    return bool(re.match(r"^\s*п\s*[/\\]?\s*ф\b", str(s or ""), re.I))


def _norm(s) -> str:
    """Нормализация имени для связывания: lower, ё→е, срез префикса п/ф, схлоп пробелов."""
    s = str(s or "").strip().lower().replace("ё", "е")
    s = _PF_PREFIX.sub("", s)  # «п/ф », «пф », «п\ф »
    s = s.replace("/", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _clean_supplier(name) -> str:
    """Имя поставщика: URL → домен; число/пусто → «—»; длинное описание → обрезаем."""
    s = str(name).strip() if name is not None else ""
    if not s or _num(s) is not None:  # пусто или чисто число — мусор из ячейки
        return "—"
    m = re.search(r"https?://([^/]+)", s)
    if m:
        return m.group(1).replace("www.", "")
    if len(s) > 40:
        return s[:40].rstrip() + "…"
    return s


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("\xa0", "").replace(" ", "")
    if s in ("", "-", "#N/A", "#REF!", "#ЗНАЧ!", "None"):
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if not v:
        return None
    for fmt in ("%d.%m.%y", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _clean_unit(v) -> str:
    s = str(v or "").strip()
    if s in ("#REF!", "#N/A", "#ЗНАЧ!", "-"):
        return ""
    return s


def _yield_unit_from_header(text) -> str:
    m = re.search(r"\(([^)]+)\)", str(text or ""))
    return m.group(1).strip(" .") if m else ""


def _portion_cost(ws) -> float | None:
    """Порционная с/с блюда из блока «Для расчёта цены продажи» → строка «Итого с/с».

    Это полная с/с ОДНОЙ порции (продукты + списания + упаковка) — достоверный источник
    с/с по блюду. В отличие от `cost_total` («Итого» col7 = с/с за весь выход карты,
    т.е. БАТЧ: напр. чай 28.84 за 4800 мл ≈ 24 чашки), эта величина не зависит от
    полноты листа «Сводная» и есть почти у всех продаваемых блюд.
    """
    # «Итого с/с» (строка-маркер) живёт в блоке ценообразования у верха листа.
    # Обычная рецептурная строка зовётся просто «Итого» — её не путаем.
    for r in range(1, min(ws.max_row, 60) + 1):
        for c in (1, 2):
            t = str(ws.cell(r, c).value or "").strip().lower()
            if "итого" in t and "с/с" in t:
                # значение порой в колонке «руб» (3), порой смещено в (4)
                return _num(ws.cell(r, 3).value) or _num(ws.cell(r, 4).value)
    return None


class _Importer:
    def __init__(self, db):
        self.db = db
        self.suppliers: dict[str, Supplier] = {}
        self.ingredients: dict[str, Ingredient] = {}  # ключ = name_norm
        self.counters = {"suppliers": 0, "ingredients": 0, "prices": 0, "ttk": 0, "lines": 0}
        # поставщиков импорт не удаляет (у них бывают файлы) — кэшируем существующих
        # по casefold-ключу (SQLite lower() не трогает кириллицу, поэтому матчим в Python)
        for s in db.query(Supplier).all():
            self.suppliers.setdefault(s.name.casefold(), s)

    # ---------- справочники ----------

    def supplier(self, name) -> Supplier:
        display = _clean_supplier(name)
        key = display.casefold()  # схлоп дублей по регистру (Ozon/OZON)
        if key not in self.suppliers:
            sup = Supplier(name=display)
            self.db.add(sup)
            self.db.flush()
            self.counters["suppliers"] += 1
            self.suppliers[key] = sup
        return self.suppliers[key]

    def ingredient(self, name, unit: str = "") -> Ingredient:
        nn = _norm(name)
        if nn not in self.ingredients:
            ing = Ingredient(name=str(name).strip(), name_norm=nn, unit=unit or "")
            self.db.add(ing)
            self.db.flush()
            self.ingredients[nn] = ing
            self.counters["ingredients"] += 1
        else:
            ing = self.ingredients[nn]
            if unit and not ing.unit:
                ing.unit = unit
        return self.ingredients[nn]

    # ---------- листы ----------

    def import_supplier_matrix(self, ws):
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name is None or str(name).strip() == "":
                continue
            sup = self.supplier(ws.cell(r, 2).value)
            ing = self.ingredient(name)
            self.db.add(
                SupplierPrice(
                    supplier_id=sup.id,
                    ingredient_id=ing.id,
                    pack_size=_num(ws.cell(r, 3).value),
                    pack_price=_num(ws.cell(r, 4).value),
                    unit_price=_num(ws.cell(r, 5).value),
                    price_date=_to_date(ws.cell(r, 6).value),
                    source="import",
                )
            )
            self.counters["prices"] += 1

    def import_product_matrix(self, ws):
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if name is None or str(name).strip() == "":
                continue
            ing = self.ingredient(name, unit=_clean_unit(ws.cell(r, 2).value))
            for col, sup_name in PRODUCT_SUPPLIER_COLS.items():
                price = _num(ws.cell(r, col).value)
                if price and price > 0:
                    sup = self.supplier(sup_name)
                    self.db.add(
                        SupplierPrice(
                            supplier_id=sup.id,
                            ingredient_id=ing.id,
                            unit_price=price,
                            source="import",
                        )
                    )
                    self.counters["prices"] += 1

    def summary_map(self, ws) -> dict[str, dict]:
        """Из «Сводной»: norm(имя блюда) → {category, price, cost_full}.

        Колонки: 1=Категория(заголовок), 2=Наименование, 3=цена, 5=cost финальный,
        11=с/с общая (с/с ПОРЦИИ: продукты+списания+упаковка). cost_full per-portion —
        достоверный источник с/с по блюду (метрика iiko относит расход к ингредиентам).
        """
        out: dict[str, dict] = {}
        current = ""
        for r in range(2, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = ws.cell(r, 2).value
            if c1 and str(c1).strip() and not (c2 and str(c2).strip()):
                current = str(c1).strip()  # строка-заголовок категории
            elif c2 and str(c2).strip() and "#" not in str(c2):
                out[_norm(c2)] = {
                    "category": current,
                    "price": _num(ws.cell(r, 3).value),
                    "cost_full": _num(ws.cell(r, 11).value),
                }
        return out

    # ---------- ТТК ----------

    @staticmethod
    def _find_header_row(ws) -> int:
        for r in range(1, min(ws.max_row, 8) + 1):
            if str(ws.cell(r, 1).value or "").strip().lower() == "ингредиент":
                return r
        return 3

    @staticmethod
    def _col_map(ws, hr: int) -> dict[str, int]:
        """Позиции колонок по шапке. Раскладка плавает: у комбо-листов нет колонок
        потерь, поэтому «с/с руб.» и «Выход» сдвинуты (5/7 вместо 7/9). Дефолты — под
        обычный 9-колоночный лист.
        """
        cols = {"gross": 2, "net": 3, "unit": 4, "cost": 7, "yield": 9}
        for c in range(1, 14):
            h = str(ws.cell(hr, c).value or "").strip().lower()
            if h == "брутто":
                cols["gross"] = c
            elif h == "нетто":
                cols["net"] = c
            elif h.startswith("ед"):
                cols["unit"] = c
            elif h.startswith("% потери"):
                cols["waste"] = c
            elif h.startswith("с/с руб"):
                cols["cost"] = c
            elif h.startswith("выход"):
                cols["yield"] = c
        return cols

    def parse_ttk(self, ws, name: str, is_semi: bool):
        """Создаёт Ttk, возвращает (ttk, raw_lines) для второго прохода."""
        hr = self._find_header_row(ws)
        cols = self._col_map(ws, hr)
        yield_unit = _yield_unit_from_header(ws.cell(hr, cols["yield"]).value)

        lines = []
        yield_qty = None
        cost_total = 0.0
        for r in range(hr + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            ls = str(label or "").strip()
            if ls == "":
                # пустая строка после состава — конец таблицы
                if lines:
                    break
                continue
            if ls.lower() == "итого":
                yield_qty = _num(ws.cell(r, cols["yield"]).value)
                cost_total = _num(ws.cell(r, cols["cost"]).value) or 0.0
                break
            lines.append(
                {
                    "raw_name": ls,
                    "gross": _num(ws.cell(r, cols["gross"]).value),
                    "net": _num(ws.cell(r, cols["net"]).value),
                    "unit": _clean_unit(ws.cell(r, cols["unit"]).value),
                    "waste_pct": _num(ws.cell(r, cols["waste"]).value) if "waste" in cols else None,
                    "cost_rub": _num(ws.cell(r, cols["cost"]).value),
                }
            )

        ttk = Ttk(
            name=name,
            name_norm=_norm(name),
            is_semi=is_semi,
            yield_qty=yield_qty,
            yield_unit=yield_unit,
            cost_total=cost_total,
        )
        self.db.add(ttk)
        self.db.flush()
        self.counters["ttk"] += 1
        return ttk, lines

    def run(self, wb) -> dict:
        # 1. прайс и номенклатура
        self.import_supplier_matrix(wb["Матрица поставщиков"])
        self.import_product_matrix(wb["Матрица продуктов"])
        summary = self.summary_map(wb["Сводная"])

        # 2. первый проход — создаём все ТТК (канон-имя = имя листа, B1 ненадёжен)
        index_all: dict[str, Ttk] = {}
        index_semi: dict[str, Ttk] = {}
        pending = []  # (ttk, lines)
        for sheet in wb.sheetnames:
            if sheet in SERVICE_SHEETS:
                continue
            ws = wb[sheet]
            name = _strip_pf(sheet) or sheet.strip()
            is_semi = sheet.strip().lower().startswith("пф")
            ttk, lines = self.parse_ttk(ws, name, is_semi)
            # данные из «Сводной» — по имени B1 либо по имени листа
            b1_norm = _norm(ws.cell(1, 2).value)
            info = summary.get(b1_norm) or summary.get(_norm(name)) or {}
            ttk.category = info.get("category", "")
            ttk.sale_price = info.get("price")
            # порционная с/с: «Итого с/с» с листа (осн.) → «Сводная» → «Итого» рецепта.
            # Последний резерв — для комбо: у них нет «Итого с/с», но «Итого» по столбцу
            # с/с уже включает упаковку и идёт за одну порцию (выход = 1 комбо).
            ttk.cost_full = _portion_cost(ws) or info.get("cost_full") or ttk.cost_total
            pending.append((ttk, lines))
            index_all.setdefault(_norm(name), ttk)
            if is_semi:
                index_semi.setdefault(_norm(name), ttk)

        # 3. второй проход — привязываем строки к сырью/п/ф
        for ttk, lines in pending:
            for ln in lines:
                key = _norm(ln["raw_name"])
                if _is_pf_ref(ln["raw_name"]):
                    child = index_semi.get(key) or index_all.get(key)
                else:
                    child = index_all.get(key)
                ingredient_id = None
                child_id = None
                if child is not None and child.id != ttk.id:
                    child_id = child.id
                else:
                    ingredient_id = self.ingredient(ln["raw_name"], ln["unit"]).id
                self.db.add(
                    TtkIngredient(
                        ttk_id=ttk.id,
                        raw_name=ln["raw_name"],
                        ingredient_id=ingredient_id,
                        child_ttk_id=child_id,
                        gross=ln["gross"],
                        net=ln["net"],
                        unit=ln["unit"],
                        waste_pct=ln["waste_pct"],
                        cost_rub=ln["cost_rub"],
                    )
                )
                self.counters["lines"] += 1

        return self.counters


def import_ttk_matrix(path: str = SEED_PATH) -> dict:
    """Идемпотентный импорт. Чистит ТТК/цены/ингредиенты, поставщиков апсёртит по имени."""
    if not os.path.exists(path):
        raise FileNotFoundError(f"seed-файл не найден: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    with SessionLocal() as db:
        # привязки блюдо↔ТТК ссылаются на ttk_id, а ТТК пересоздаются с новыми id —
        # запоминаем нормализованное имя цели, чтобы перелинковать после импорта
        old_target = {
            m.id: (m.ttk.name_norm if m.ttk else None) for m in db.query(DishMapping).all()
        }
        # торговые марки заданы вручную на SupplierPrice, а прайс пересоздаётся — сохраняем
        # по ключу (имя поставщика casefold, name_norm ингредиента), затем восстанавливаем
        brand_map: dict[tuple[str, str], str] = {}
        for sp in db.query(SupplierPrice).filter(SupplierPrice.brand != "").all():
            sup = db.get(Supplier, sp.supplier_id)
            ing = db.get(Ingredient, sp.ingredient_id)
            if sup and ing:
                brand_map[(sup.name.casefold(), ing.name_norm)] = sp.brand

        # очистка прежнего импорта (поставщиков и привязки НЕ трогаем)
        db.query(TtkIngredient).delete()
        db.query(Ttk).delete()
        db.query(SupplierPrice).delete()
        db.query(Ingredient).delete()
        db.flush()

        counters = _Importer(db).run(wb)

        # восстановление торговых марок на новых строках прайса
        if brand_map:
            for sp in db.query(SupplierPrice).all():
                sup = db.get(Supplier, sp.supplier_id)
                ing = db.get(Ingredient, sp.ingredient_id)
                if sup and ing:
                    brand = brand_map.get((sup.name.casefold(), ing.name_norm))
                    if brand:
                        sp.brand = brand

        # перелинковка привязок по имени ТТК (предпочитаем НЕ-п/ф — is_semi False первыми)
        new_by_norm: dict[str, int] = {}
        for t in db.query(Ttk).order_by(Ttk.is_semi).all():
            new_by_norm.setdefault(t.name_norm, t.id)
        for m in db.query(DishMapping).all():
            nn = old_target.get(m.id)
            if nn and nn in new_by_norm:
                m.ttk_id = new_by_norm[nn]

        db.add(SyncLog(sync_type="import_ttk", status="ok", message=str(counters)))
        db.commit()

    logger.info("Импорт ТТК завершён: %s", counters)
    return counters
